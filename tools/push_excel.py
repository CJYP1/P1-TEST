#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把《区域划分 EB_NB_MA》Excel 的排程时间推入系统。

写入内容:
  1. 每条月度工作安排的 计划开始/计划结束(ps/pf)← Excel 逐条活动的真实日期;
  2. data-csv/fixed/zone-schedule.csv ← 表1/表2 的区域施工时间(大区/楼层/分区起止);
  3. data-csv/fixed/conflicts.csv     ← 冲突数据(数量不一致/月份不一致/完成率不符/
                                        缺排程/系统无对应项), 供人工验证 —— 冲突项只标记不覆盖。
不改动任何 计划量p/实际量d/完成率pct —— 月度显示与后续实际量更新不受影响。
运行: python tools/push_excel.py   (Excel 在 data-csv/source/ 下)
"""
import json, re, csv, datetime, calendar
from pathlib import Path
from collections import defaultdict
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XL = ROOT/'data-csv'/'source'/'RWSP1_区域划分_EB_NB_MA1.xlsx'

# ---------- load ZP ----------
t = (ROOT/'zp-data.global.js').read_text(encoding='utf-8')
lines = t.split('\n')
gi = next(i for i, l in enumerate(lines) if l.startswith('window.__RWS.ZP'))
ZP = json.loads(lines[gi][lines[gi].index('{'):lines[gi].rindex('}')+1].rstrip(';'))
zp_zones = {z for b in ZP['buildings'] for z in b['zones']} | set(ZP['plan'])

def norm(z): return re.sub(r'[\s\-()]', '', str(z).strip()).upper()
zp_by_norm = {}
for z in zp_zones: zp_by_norm.setdefault(norm(z), z)

# —— 分区对照表(项目文件 zone-xref.csv)= 命名基准: 同组名称视为同一区域 ——
xref_group = {}          # norm(name) -> group_id
with open(ROOT/'data-csv'/'fixed'/'zone-xref.csv', encoding='utf-8-sig') as f:
    rd = csv.reader(f); next(rd)
    for row in rd:
        if len(row) >= 5 and row[0].strip():
            xref_group[norm(row[4])] = row[0].strip()
def group_of(name):
    n = norm(name)
    for c in (n, 'POD'+n, n[3:] if n.startswith('POD') else None, 'P'+n, n[1:] if n.startswith('P') else None):
        if c and c in xref_group: return xref_group[c]
    return None
zp_group = {z: group_of(z) for z in zp_zones}

def map_zone(bld, lvl, zone):
    """Excel 分区名 → 系统(ZP)分区名"""
    raw = str(zone).strip(); n = norm(raw)
    cands = [n]
    if bld=='EB' and lvl=='B2' and not n.startswith('EB'): cands.append('EB2'+n)      # '3.3'→EB2-3.3
    if bld=='EB' and lvl=='B1' and re.match(r'^S\d', n): cands.append('EB1'+n)        # 'S2.1'→EB1S2.1
    if bld=='EB' and re.match(r'^B[23]\.', n): cands.append('B'+n)                    # 'B2.1'→B-2.1
    if re.match(r'^\d', n): cands.append('POD'+n)                                     # 塔楼 → POD*
    if n=='EB23.4+3.5': cands.append('EB23.3+3.5')
    if n=='TOPSLAB': cands.append('TOPSLAB')
    for c in cands:
        if c in zp_by_norm: return zp_by_norm[c]
    return None

# ---------- 工作项类型识别 ----------
def type_of_label(w):
    w = str(w)
    if 'Steel Beam' in w: return 'sbeam'
    if 'Main Beam' in w: return 'mbeam'
    if 'Column' in w: return 'col'
    if 'Corbel' in w: return 'corbel'
    if 'Core Wall' in w: return 'core'
    if 'Lift/Stair' in w or 'Lift / Stair' in w: return 'ls'
    if 'Top Slab' in w or 'Slab' in w: return 'slab'
    if 'Demoli' in w: return 'demo'
    if 'Pilecap' in w or 'Pile Cap' in w: return 'pile'
    if 'Excavat' in w: return 'exc'
    if 'Piling' in w: return 'con'
    return None
def floor_of_label(w):
    w = str(w)
    m = re.search(r'L(\d)\s*-\s*L(\d)', w)
    if m: return f'L{max(int(m.group(1)),int(m.group(2)))}'
    if 'B1M' in w: return 'B1M'
    if re.search(r'\bB2\b', w): return 'B2'
    if re.search(r'\bB1\b', w): return 'B1'
    m = re.search(r'\bL(\d)\b', w)
    return f'L{m.group(1)}' if m else '其他'

# ---------- read excel 逐条 ----------
wb = openpyxl.load_workbook(XL, data_only=True)
def to_date(v):
    if hasattr(v, 'date'): return v.date() if not isinstance(v, datetime.date) else v
    return datetime.datetime.strptime(str(v).strip()[:10], '%Y-%m-%d').date()

TRADE2TYPE = {'COL':'col','MBR':'mbeam','STB':'sbeam','SLB':'slab','EXC':'exc','DEM':'demo',
              'PIL':'pile','CON':'con','RCW':'slab','RCWDT':'slab'}
TOWER = re.compile(r'^\d')
SHIFT = {'L1':'L2','L2':'L3','L3':'L4','L4':'L5'}
groups = defaultdict(list)     # (zpzone, floor, type) -> [ {qty,unit,start,finish,src} ]
xgroups = defaultdict(list)    # ('G'+对照组, floor, type) -> 同上(命名基准兜底匹配)
unmapped = set()
for sn, bld in [('3_EB_逐条','EB'), ('4_NB_逐条','NB'), ('5_MA_逐条','MA')]:
    ws = wb[sn]
    for r in range(4, ws.max_row+1):
        lvl, zone, trade, _aid, name, qty, s0, f0 = (ws.cell(r, c).value for c in range(1, 9))
        if not zone or not lvl: continue
        lvl = str(lvl).strip(); trade = str(trade).strip(); name = str(name or '')
        if trade in ('EAR','CYC','MEP'): continue
        typ = TRADE2TYPE.get(trade)
        if trade in ('CRW','WAL'): typ = 'core' if 'Core Wall' in name else 'ls'
        if typ is None: continue
        try: sd, fd = to_date(s0), to_date(f0)
        except Exception: continue
        zp = map_zone(bld, lvl, str(zone))
        if zp is None:
            unmapped.add(f'{bld}/{lvl}/{zone}'); continue
        pairs = [(float(m.group(1).replace(',','')), {'㎡':'m²'}.get(m.group(2), m.group(2)))
                 for m in re.finditer(r'([\d,.]+)\s*(m²|㎡|m³|nos)', str(qty))]
        # 主量
        subs = []
        prim = {'col':'nos','mbeam':'nos','sbeam':'nos','core':'nos','ls':'nos','pile':'nos','con':'nos',
                'slab':'m²','exc':'m³','demo':None}[typ]
        for v, u in pairs:
            if typ=='demo' or (prim and u==prim): subs.append((typ, v))
            elif u=='nos' and ('Pilecap' in name or 'pilecap' in name or 'Piling' in name): subs.append(('pile', v))
            elif u=='nos' and 'Beam' in name: subs.append(('mbeam', v))
        if not subs: subs = [(typ, None)]
        for styp, v in subs:
            fl = lvl
            if styp in ('col','mbeam','sbeam','core','ls','corbel') and lvl in SHIFT and TOWER.match(norm(zone)):
                fl = SHIFT[lvl]
            rec = {'qty': v, 'start': sd, 'finish': fd, 'src': f'{sn} r{r}'}
            groups[(zp, fl, styp)].append(rec)
            gid = group_of(zp) or group_of(zone)
            if gid: xgroups[('G'+gid, fl, styp)].append(rec)

# ---------- match & push dates ----------
conflicts = []
n_dated = 0; n_items = 0; n_nosched = 0
MONNUM = {'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,'July':7,'August':8,
          'September':9,'October':10,'November':11,'December':12}
def mon_range(label):
    m = re.match(r'([A-Za-z]+)\s+(\d{4})', label); y, mo = int(m.group(2)), MONNUM[m.group(1)]
    return (datetime.date(y, mo, 1), datetime.date(y, mo, calendar.monthrange(y, mo)[1]))

zp_group_sum = defaultdict(float)
for zn, ms in ZP['plan'].items():
    for mon, items in ms.items():
        for idx, it in enumerate(items):
            typ = type_of_label(it.get('w')); fl = floor_of_label(it.get('w'))
            n_items += 1
            if typ is None: continue
            key = (zn, fl, typ)
            if it.get('p') is not None: zp_group_sum[key] += it.get('p') or 0
            rows = groups.get(key)
            src = '精确'
            if not rows:
                gid = zp_group.get(zn)
                if gid: rows = xgroups.get(('G'+gid, fl, typ)); src = '对照组'
            it.pop('tnote', None)
            if not rows:
                if it.get('p'):
                    it['tnote'] = '无Excel排程'
                    n_nosched += 1
                continue
            m0, m1 = mon_range(mon)
            ov = [r for r in rows if r['start'] <= m1 and r['finish'] >= m0]
            use = ov if ov else rows
            it['ps'] = min(r['start'] for r in use).isoformat()
            it['pf'] = max(r['finish'] for r in use).isoformat()
            n_dated += 1
            if not ov:
                it['tnote'] = '⚠月份与Excel排程不符'
                conflicts.append(['月份不一致(日期已推入)', zn, it.get('w'), mon,
                                  f"系统在{mon}有计划{it.get('p')}{it.get('u','')}",
                                  f"Excel排程 {it['ps']}~{it['pf']}", '日期已按Excel写入, 月份归属请人工核对'])
            # 完成率核对
            p, d, pct = it.get('p'), it.get('d'), it.get('pct')
            if p and d is not None and pct is not None:
                calc = round(d / p * 100)
                if abs(calc - pct) > 1:
                    conflicts.append(['完成率不符', zn, it.get('w'), mon,
                                      f'存 pct={pct}%', f'按 d/p 计算={calc}% (d={d}, p={p})',
                                      '完成率与 实际/计划 不一致 — 请人工确认哪个正确'])

# 数量合计核对(组级)
for key, rows in sorted(groups.items()):
    zn, fl, typ = key
    ex_sum = sum(r['qty'] for r in rows if r['qty'] is not None)
    zp_sum = zp_group_sum.get(key)
    if zp_sum is None or zp_sum == 0:
        if ex_sum:
            conflicts.append(['系统无对应项', zn, f'{fl} {typ}', '',
                              '系统无此组计划', f'Excel 合计 {ex_sum:g}',
                              'Excel 有排程但系统月度网格无此工作项 — 请人工确认是否需补录'])
        continue
    if ex_sum and abs(ex_sum - zp_sum) > max(1, zp_sum * 0.02):
        conflicts.append(['数量不一致', zn, f'{fl} {typ}', '(合计)',
                          f'系统合计 {zp_sum:g}', f'Excel 合计 {ex_sum:g}',
                          '两边总量不同 — 冲突未覆盖, 请人工裁定后改 work CSV 或 Excel'])

# ---------- zone-schedule.csv (表1/表2 区域施工时间) ----------
sched = []
for sn, area_col in [('1_区域边界_EB_MA', True), ('2_区域边界_NB', False)]:
    ws = wb[sn]
    for r in range(5, ws.max_row+1):
        b, l, z, full, rule, s0, f0, dur, wl, _x, span, _f = (ws.cell(r, c).value for c in range(1, 13))
        if not z: continue
        bld = str(b).strip() if area_col else 'NB'
        if not area_col: bld = 'NB'
        elif bld not in ('EB','MA'): continue
        zp = map_zone(bld, str(l).strip(), str(z).strip())
        sched.append([bld, str(l).strip(), str(z).strip(), zp or '(未匹配)', str(full or ''),
                      str(s0 or '')[:10], str(f0 or '')[:10], str(dur or ''), str(wl or ''), str(span or '')])
with open(ROOT/'data-csv'/'fixed'/'zone-schedule.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(['大区','楼层','分区(Excel)','系统分区','小区全称','区域开始','区域结束','时长','工作量','全区跨度'])
    w.writerows(sched)

# ---------- write ----------
lines[gi] = 'window.__RWS.ZP = ' + json.dumps(ZP, ensure_ascii=False, separators=(',', ':')) + ';'
(ROOT/'zp-data.global.js').write_text('\n'.join(lines), encoding='utf-8')
with open(ROOT/'data-csv'/'fixed'/'conflicts.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(['冲突类型','分区','工作项/组','月份','系统值','Excel值','处理建议'])
    w.writerows(conflicts)
from collections import Counter
print(f'✔ 已为 {n_dated}/{n_items} 条工作安排写入真实计划日期 (p/d/pct 未动); 无Excel排程 {n_nosched} 条(标记在work CSV)')
print(f'区域施工时间表: {len(sched)} 行 → data-csv/fixed/zone-schedule.csv')
print(f'冲突清单: {len(conflicts)} 条 → data-csv/fixed/conflicts.csv  ' + str(dict(Counter(c[0] for c in conflicts))))
if unmapped: print('未匹配的 Excel 分区:', sorted(unmapped))
